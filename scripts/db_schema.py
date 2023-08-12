import openpyxl
import sys

# Load the Excel file
excel_file = sys.argv[1]
workbook = openpyxl.load_workbook(excel_file)

tables = {}
# Iterate through all sheets
for sheet_name in workbook.sheetnames:
    if sheet_name == "Entities list":
        continue
    sheet = workbook[sheet_name]

    table_name = sheet_name.split("(")[-1].split(")")[0].lower()

    columns = []
    candidate_primary_keys = set()
    foreign_keys = []
    for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
        if i > 0:
            col_name, _, _, col_type, *rest = row
            # print(col_name, col_type)
            if col_type == "Uniqueidentifier":
                candidate_primary_keys.add(col_name)
            elif col_type == "Lookup":
                reference_table = row[8].split("\n")[1]
                foreign_keys.append((col_name, reference_table))
            columns.append(col_name)

    column_defs = []

    if table_name + "id" in candidate_primary_keys:
        primary_key = table_name + "id"
    elif len(candidate_primary_keys) == 1:
        primary_key = candidate_primary_keys.pop()
    else:
        primary_key = None

    for column in columns:
        if column == primary_key:
            column_defs.append(f"{column} text primary key")
        else:
            column_defs.append(f"{column} text")

    tables[table_name] = {"defs": column_defs, "fks": foreign_keys, "pk": primary_key}


# Close the workbook
workbook.close()

for name, definitions in tables.items():
    foreign_keys = [
        f"FOREIGN KEY ({col_name}) REFERENCES {table_name}({tables[table_name]['pk']})"
        for col_name, table_name in definitions["fks"]
        if table_name in tables
    ]
    print(f"create table {name} ({', '.join(definitions['defs'] + foreign_keys)});")
